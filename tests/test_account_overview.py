import re
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import expect

from utils.excel_writer import write_account_overview_to_excel


def test_accounts_overview_page_is_displayed(account_overview_page):
    """Validation test: verify heading, table visibility, headers, and navigation."""
    account_overview_page.assert_navigation_url()
    account_overview_page.assert_accounts_overview_displayed()
    account_overview_page.assert_table_headers()
    account_overview_page.assert_balance_note_displayed()


def test_accounts_are_listed_with_valid_values(account_overview_page):
    """Positive test: verify account rows contain valid account and amount values."""
    assert account_overview_page.get_total_accounts_count() > 0
    account_overview_page.assert_account_rows_are_valid()


def test_total_balance_is_displayed_in_currency_format(account_overview_page):
    """Validation test: verify total balance is displayed with a valid format."""
    account_overview_page.assert_total_balance_format()
    account_overview_page.assert_total_balance_matches_account_rows()


def test_first_account_link_is_enabled(account_overview_page):
    """UI validation test: verify account links are visible and enabled."""
    first_account_link = account_overview_page.page.locator(
        account_overview_page.account_links
    ).first

    expect(first_account_link).to_be_visible()
    expect(first_account_link).to_be_enabled()


def test_user_can_open_account_activity_page(account_overview_page):
    """UI interaction test: verify clicking an account opens its activity page."""
    account_id = account_overview_page.open_account_by_index(0)

    expect(account_overview_page.page).to_have_url(
        re.compile(rf".*/activity\.htm\?id={account_id}$")
    )


def test_export_account_overview_table_to_excel(account_overview_page):
    """Positive test: export account overview web table data to Excel."""
    account_rows = account_overview_page.get_account_overview_rows()
    summary = account_overview_page.get_account_summary()
    output_file = Path("reports/account_overview.xlsx")

    exported_file = write_account_overview_to_excel(
        account_rows=account_rows,
        summary=summary,
        output_path=output_file,
    )

    assert exported_file.exists()

    workbook = load_workbook(exported_file)
    details_sheet = workbook["Account Details"]
    summary_sheet = workbook["Summary"]

    assert details_sheet.max_row == len(account_rows) + 1
    assert details_sheet.max_column == 4
    assert summary_sheet["A1"].value == "Metric"
    assert summary_sheet["B1"].value == "Value"
    assert details_sheet["A2"].value == account_rows[0]["account_number"]
